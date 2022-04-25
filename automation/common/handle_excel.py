# coding=utf-8

"""
@author:lyle.hong
@date:2022/4/15 16:59
"""

import openpyxl


class HandleExcel:
    """操作Excel"""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self):
        """
        读取Excel文件内容
        :return:
        """
        workbook = openpyxl.load_workbook(self.file_name)
        sh = workbook[self.sheet_name]
        res = list(sh.rows)
        # 取到首行数据
        title = [i.value for i in res[0]]
        list_data = []
        for i in res[1:]:
            data = [j.value for j in i]
            list_data.append(dict(zip(title, data)))
        return list_data

    def write_data(self, row, column, value):
        """
        写入Excel
        :return:
        """
        workbook = openpyxl.load_workbook(self.file_name)
        sh = workbook[self.sheet_name]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.file_name)


if __name__ == '__main__':
    a = HandleExcel(r"D:\automation\datas\官网注册接口用例.xlsx", "Sheet1")
    print(a.read_data())
    a.write_data(8, 8, "sssss")
